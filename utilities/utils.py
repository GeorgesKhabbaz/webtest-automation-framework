"""
utils.py
Centralized helper functions and utilities used across the framework.
Includes logging, screenshot capture, Excel utilities, and file/folder helpers.
"""

import logging
import os
from datetime import datetime
from openpyxl import load_workbook

# ----------------------------
# Logger setup with log files
# ----------------------------

# Ensure logs directory exists
LOGS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "logs")
os.makedirs(LOGS_DIR, exist_ok=True)

def setup_logger(name="framework_logger", log_level=logging.INFO):
    """
    Creates and configures a logger.
    Each test run generates a timestamped log file in the logs/ folder.

    Args:
        name (str): Name of the logger
        log_level (int): Logging level (default: INFO)

    Returns:
        logging.Logger: Configured logger instance
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    log_file = os.path.join(LOGS_DIR, f"execution_{timestamp}.log")

    logger = logging.getLogger(name)
    logger.setLevel(log_level)

    # Prevent duplicate handlers if called multiple times
    if not logger.handlers:
        # File handler (log file)
        file_handler = logging.FileHandler(log_file, encoding="utf-8")
        # Console handler (prints to terminal)
        console_handler = logging.StreamHandler()

        formatter = logging.Formatter(
            "%(asctime)s | %(levelname)s | %(name)s | %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S"
        )

        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)

        logger.addHandler(file_handler)
        logger.addHandler(console_handler)

    return logger

# Example usage:
# logger = setup_logger()
# logger.info("Framework started")


# ----------------------------
# Screenshot capture
# ----------------------------
def take_screenshot(driver, file_name="screenshot"):
    """
    Captures a screenshot of the current browser window.

    Args:
        driver (WebDriver): Selenium WebDriver instance
        file_name (str): Base name for the screenshot file

    Returns:
        str: Full path to the saved screenshot
    """
    screenshots_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "screenshots")
    os.makedirs(screenshots_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(screenshots_dir, f"{file_name}_{timestamp}.png")
    driver.save_screenshot(path)
    return path


# ----------------------------
# Excel utilities
# ----------------------------
def read_excel(file_path, sheet_name):
    """
    Reads data from an Excel file (.xlsx) and returns it as a list of dictionaries.

    Args:
        file_path (str): Path to the Excel file
        sheet_name (str): Name of the sheet to read

    Returns:
        list[dict]: List of rows with column headers as keys
    """
    workbook = load_workbook(filename=file_path)
    sheet = workbook[sheet_name]
    data = []
    headers = [cell.value for cell in sheet[1]]  # First row = headers

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

# Example usage:
# test_data = read_excel("testdata/test_data.xlsx", "login")
# print(test_data)


# ----------------------------
# File / folder utilities
# ----------------------------
def ensure_folder_exists(folder_path):
    """
    Creates a folder if it does not exist.

    Args:
        folder_path (str): Path to the folder
    """
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

# Example usage:
# ensure_folder_exists("reports")
